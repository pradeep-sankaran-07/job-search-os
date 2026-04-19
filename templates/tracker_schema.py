"""Create a fresh tracker.xlsx with the Job Search OS schema.

Columns (in order): ID | Date Found | Job Title | Company | Location | Source |
                    Fit Rating | Recommendation | Key Notes | Cover Letter Text |
                    Apply By | Status

Sheets: Jobs (active), Archive (Applied, Will not apply, Not a real job).

Status dropdown values: New, Applied, Will not apply, Not a real job.

Usage:
    python tracker_schema.py /path/to/tracker.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

COLUMNS = [
    "ID",
    "Date Found",
    "Job Title",
    "Company",
    "Location",
    "Source",
    "Fit Rating",
    "Recommendation",
    "Key Notes",
    "Cover Letter Text",
    "Apply By",
    "Status",
]

COLUMN_WIDTHS = {
    "A": 6,    # ID
    "B": 12,   # Date Found
    "C": 28,   # Job Title
    "D": 22,   # Company
    "E": 22,   # Location
    "F": 50,   # Source (URL)
    "G": 12,   # Fit Rating
    "H": 16,   # Recommendation
    "I": 60,   # Key Notes
    "J": 60,   # Cover Letter Text
    "K": 12,   # Apply By
    "L": 16,   # Status
}

STATUS_VALUES = '"New,Applied,Will not apply,Not a real job"'


def build_tracker(output_path: Path) -> None:
    """Create (or overwrite) a tracker.xlsx at output_path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    # Default sheet becomes Jobs.
    ws_jobs = wb.active
    ws_jobs.title = "Jobs"

    ws_archive = wb.create_sheet("Archive")

    for ws in (ws_jobs, ws_archive):
        for col_idx, header in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Freeze header row + first four columns (ID, Date, Title, Company).
        ws.freeze_panes = "E2"

        for col_letter, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

    # Status dropdown on Jobs sheet, rows 2-1000.
    dv = DataValidation(type="list", formula1=STATUS_VALUES, allow_blank=True, showDropDown=False)
    dv.add("L2:L1000")
    ws_jobs.add_data_validation(dv)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("usage: tracker_schema.py <output_path>")
        return 2
    out = Path(argv[1]).expanduser().resolve()
    build_tracker(out)
    print(f"Created tracker: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
